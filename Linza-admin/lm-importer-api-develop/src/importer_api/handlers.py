import sys
from os.path import dirname, abspath
# Hack to import shared modules from parent directory
sys.path.append(dirname(dirname(abspath(__file__))))
from config import Config
from importer_api.schemas import HealthSchema, DataSourceSchema, DataSourceItemSchema, \
DataSourceItemSearchSchema, SocialNetworkCSVSchema, IntegrumSchema, SocialIntegrumSchema, \
DataSourcePostSchema, DataSourcePutSchema, JobSchema, DataSourceItemPutSchema, DataSourcePatchSchema, \
DataSourceItemPatchSchema, DataSourcesItemsByUidsSchema, DataSourceItemConfigSchema
from models import Health, DataSource, DataSourceItem, SocialNetworkCSV, Integrum, \
    SocialIntegrum, Job, filestore, datastore
from producer import Producer
from shared import EventsSender, RabbitConfigMaker, get_logger
from collections import namedtuple, OrderedDict
from openpyxl import load_workbook, Workbook
import datetime
from datetime import datetime
import io
import peewee
import json
from tornado.escape import json_decode
import tornado.web
from uuid import uuid4
import requests
import re
import uuid
import logging
import os
from tornado import escape
import csv
from transliterate import translit
from urllib.parse import quote, urlsplit, urlunsplit
from iteration_utilities import unique_everseen
import socket


def is_uuid4(string):
    return re.match(r"([0-9a-f]{8})-([0-9a-f]{4})-([0-9a-f]{4})-([0-9a-f]{4})-([0-9a-f]{12})$", string)



logger = get_logger()


class BaseHandler(tornado.web.RequestHandler):

    def set_default_headers(self):
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Headers", "x-requested-with, "
                                                        "access-control-allow-origin, "
                                                        "content-type, "
                                                        "authorization")

    def options(self, **kwargs):
        self.set_header('Access-Control-Allow-Methods', 'OPTIONS')
        self.set_status(204)
        self.finish()


class DataRequestHandler(BaseHandler):

    def prepare(self):
        if datastore.is_closed():
            datastore.connect()
        return super(DataRequestHandler, self).prepare()

    def on_finish(self):
        if not datastore.is_closed():
            datastore.close()
        return super(DataRequestHandler, self).on_finish()


class HealthHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self):
        """Returns API health status.
        ---
        description: Returns API health status.
        responses:
            200:
                description: Success
                schema:
                    type: object
                    properties:
                        statusCode:
                            type: integer
                            description: StatusCode
                        description:
                            type: string
                            description: description
                        errorCode:
                            type: string
                            description: errorCode
        """
        healthsSchema = HealthSchema()
        healths = Health.select()

        result, errors = healthsSchema.dump(healths, many=True)
        self.set_status(200)
        response = {"statusCode": 200, "description": result[0]["name"],
                    "errorCode": None}
        response = {"response": response}
        self.write(response)    


class DataSourceItemHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"]

    def options(self, item_id):
        self.set_header('Access-Control-Allow-Methods', 'GET, POST, PUT, PATCH, DELETE, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self, item_id):        
        """Get Item information
        ---
        description: Get Item information
        parameters:        
        -   name: item_id
            in: path
            description: item_id
            required: true
            type: string        
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                description:
                                    type: string
                                    description: description
                                data:
                                    type: object
                                    description: result
        """
        if is_uuid4(item_id):
            item = None
            if DataSourceItem.select().where(DataSourceItem.id == item_id).exists():
                try:
                    item = DataSourceItem.select().where(DataSourceItem.id == item_id).get()
                except Exception as e:
                    logger.error("Item {} does not exist:{}".format(item_id, e))
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Item does not exist:{}".format(e)}
                    self.write(response)
                if item:
                    result, errors = DataSourceItemSchema().dump(item)
                    if errors:
                        logger.error("Error serializing DataSourceItem: {}", errors)
                    else:
                        self.set_status(200)
                        response = {"statusCode": 200, "description": "Success", "data": result}
                        self.write(response)
            else:
                logger.error("Item {} does not exist".format(item_id))
                self.set_status(404)
                response = {"statusCode": 404, "description": "Item {} not found".format(item_id)}
                self.write(response)
        else:
            logger.error("Item {} does not exist".format(item_id))
            self.set_status(404)
            response = {"statusCode": 404, "description": "Invalid uuid4 provided"}
            self.write(response)

    def post(self, item_id):        
        """Creates a new DataSourceItem.
        ---
        description: Creates a new DataSourceItem.
        parameters:        
        -   name: item_id
            in: path
            description: item_id
            required: true
            type: string
        responses:
            201:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        data:
                            type: string
                            description: result
                        description:
                            type: string
                            description: description
        """
        raw_body = None
        if is_uuid4(item_id):
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in POST body: {}".format(e))
                self.set_status(400)
                logger.error("Error in request body: {}".format(e))
                response = {"statusCode": 400, "description": "Error in request body: {}".format(e),
                            "errorCode": None}
                response = {"response": response}
                self.write(response)
                return
            if raw_body:
                message_sent = False
                logger.info("raw body: {}".format(raw_body))
                dsi_schema = DataSourceItemSchema()
                datasourceitem, errors = dsi_schema.load(raw_body)
                if len(errors) == 0:               
                    if not DataSourceItem.select().where(DataSourceItem.id == datasourceitem["id"]).exists():
                        with datastore.atomic() as tx:
                            item_saved = False
                            try:
                                new_dataSourceItem = DataSourceItem(**datasourceitem)
                                new_dataSourceItem.updated = datetime.utcnow()
                                new_dataSourceItem.save(force_insert=True)
                                logger.info("Item saved")
                                item_saved = True
                            except Exception as e:
                                logger.error(("Error saving new dataSourceItem: {}".format(e)))
                            logger.info("new_dataSourceItem: {}".format(new_dataSourceItem))
                            # new_item = dsi_schema.load(new_dataSourceItem)[0]
                            result, error = dsi_schema.dump(new_dataSourceItem)
                            """v.0.9
                            DataSource created Event sending to Bus 
                            """
                            # ds_msg = prepare_0_9(item=new_dataSourceItem)
                            if item_saved:
                                producer = Producer(exchange=Config.IMPORTER_DATASOURCE_CREATED_EXCHANGE_0_9,
                                                    type_='direct',
                                                    queue=Config.IMPORTER_DATASOURCE_CREATED_QUEUE_0_9,
                                                    routing_key=Config.IMPORTER_DATASOURCE_CREATED_ROUTING_KEY_0_9,
                                                    config=RabbitConfigMaker.main_rabbit_config()
                                                    )
                                try:
                                    EventsSender.send_0_9_event(item=new_dataSourceItem, producer=producer)
                                    message_sent = True
                                except Exception as exc:
                                    logger.error("Error sending DataSource Created Event: {}".format(exc))
                            """
                            End of Event
                            """
                            if message_sent:
                                tx.commit()
                                response = {"statusCode": 201,
                                            "description": "Success",
                                            "data": result}
                                self.set_status(201)
                                self.set_header('Location', "http://" + Config.IMPORTER_API_HOST + ":" + str(Config.IMPORTER_API_PORT) +
                                                self.reverse_url("api:datasourceitem", new_dataSourceItem.id))
                                self.write(response)
                            else:
                                tx.rollback()
                                logger.info("Transaction rolled back")
                                self.set_status(500)
                                response = {"statusCode": 500,
                                            "description": "Could not send DataSource Created 0.9 event. Transaction "
                                                            "rolled back"}
                                self.write(response)
                    else:
                        self.set_status(400)
                        logger.error("DataSourceItemUid {} already exists".format(datasourceitem["id"]))
                        response = {"statusCode": 400,
                                    "description": "DataSourceItemUid {} already exists".format(datasourceitem["id"])}
                        self.write(response)                
                else:
                    self.set_status(400)
                    logger.error("Error creating DataSourceItem: {}".format(errors))
                    response = {"statusCode": 400, "description": "Error creating DataSourceItem: {}".format(errors)}
                    self.write(response)
        else:
            logger.error("Item {} does not exist".format(item_id))
            self.set_status(404)
            response = {"statusCode": 404, "description": "Invalid uuid4 provided"}
            self.write(response)


    def patch(self, item_id):        
        """Patch dataSourceItem
        ---
        description: Patch dataSourceItem
        parameters:        
        -   name: item_id
            in: path
            description: item_id
            required: true
            type: string        
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                description:
                                    type: string
                                    description: description
                                data:
                                    type: object
                                    description: result
        """
        if is_uuid4(item_id):
            allowed_fields = ["title", "description", "iconUri", "type", "url", "updateInterval", 
                            "isActive", "isAutoCreated", "config", "account", "timezone"]
            raw_body = None
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in request body: {}".format(e))
                self.set_status(400)
                response = {"statusCode": 400, "description": "Errors in request body: {}".format(e)}
                self.write(response)
                return
            if raw_body:
                dsi_patch_schema = DataSourceItemPatchSchema()
                body, errors = dsi_patch_schema.load(raw_body)
                error_list = []
                fields_list = []
                for key, value in raw_body.items():
                    if key in allowed_fields:
                        pass
                    else:
                        logger.error("Field {} not found or not allowed".format(key))
                        d = {key: value}
                        error_list.append(d)
                if all([len(error_list) == 0, len(errors) == 0]):
                    logger.info("No errors in request body")
                    logger.info("body: {}".format(body))
                    try:
                        dsi = DataSourceItem.select().where(DataSourceItem.id == item_id).get()
                    except peewee.DoesNotExist as e:
                        logger.error("Couldn't find DataSourceItem: {}".format(e))
                        self.set_status(404)
                        response = {"statusCode": 404, "description": "Not found. {}".format(e)}
                        self.write(response)
                        return
                    item_updated = False
                    with datastore.atomic() as tx:
                        try:
                            q = DataSourceItem.update(**body).where(DataSourceItem.id == item_id)
                            q.execute()
                            q = DataSourceItem.update({DataSourceItem.modified: datetime.utcnow()}).where(
                                DataSourceItem.id == item_id)
                            q.execute()
                            item_updated = True
                        except Exception as e:
                            logger.error("Error updating item: {}".format(e))
                        if item_updated:
                            logger.info("DataSourceItem updated")

                            dsi_schema = DataSourceItemSchema()
                            updated_dsi = DataSourceItem.select().where(DataSourceItem.id == item_id).get()
                            """send 0.9 event"""
                            # data = prepare_0_9(updated_dsi)
                            sent = False
                            producer = Producer(exchange=Config.IMPORTER_DATASOURCE_UPDATED_EXCHANGE_0_9,
                                                type_='direct',
                                                queue=Config.IMPORTER_DATASOURCE_UPDATED_QUEUE_0_9,
                                                routing_key=Config.IMPORTER_DATASOURCE_UPDATED_ROUTING_KEY_0_9,
                                                config=RabbitConfigMaker.main_rabbit_config())
                            try:
                                EventsSender.send_0_9_event(item=updated_dsi, producer=producer)
                                sent = True
                            except Exception as e:
                                logger.error("Could not send 0.9 updated event: {}".format(e))
                            if sent:
                                tx.commit()
                                self.set_status(200)
                                response = {"statusCode": 200, "description": "Success", "data": dsi_schema.dump(updated_dsi)[0]}
                                self.write(response)
                            else:
                                tx.rollback()
                                logger.error("Error, transaction rolled back")
                                self.set_status(500)
                                response = {"statusCode": 500, "description": "Transaction rolled back for connection reason"}
                                self.write(response)
                else:
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Bad request. {}, {}".format(errors, error_list)}
                    self.write(response)
        else:
            self.set_status(400)
            response = {"statusCode": 400, "description": "Bad request: {}".format(item_id)}
            self.write(response)

    def put(self, item_id):        
        """UPDATE DatasourceItem.
        ---
        description: UPDATE DatasourceItem.
        parameters:        
        -   name: item_id
            in: path
            description: item_id
            required: true
            type: string
        -   name: title
            in: body
            description: title
            required: true
            type: string
        -   name: iconUri
            in: body
            description: iconUri
            required: true
            type: string
        -   name: type
            in: body
            description: type
            required: true
            type: string
        -   name: url
            in: body
            description: url
            required: true
            type: string
        -   name: updateInterval
            in: body
            description: updateInterval
            required: false
            type: integer
        -   name: dataSource
            in: body
            description: item_id
            required: true
            type: object
        -   name: isActive
            in: body
            description: isActive
            required: true
            type: boolean
        -   name: isAutoCreated
            in: body
            description: isAutoCreated
            required: false
            type: boolean
        -   name: config
            in: body
            description: config
            required: false
            type: object
        -   name: account
            in: body
            description: account
            required: false
            type: string
        -   name: timezone
            in: body
            description: timezone
            required: false
            type: string
        -   name: updated
            in: body
            description: updated
            required: false
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                description:
                                    type: string
                                    description: description
                                data:
                                    type: object
                                    description: result
        """
        raw_body = None
        if is_uuid4(item_id):
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in request body: {}".format(e))
                self.set_status(404)
                response = {"statusCode": 404, "description": "Error in request body: {}".format(e)}
                self.write(response)
                return
            if raw_body:                
                dsi_schema = DataSourceItemPutSchema()
                dsi, errors = dsi_schema.load(raw_body)
                if len(errors) == 0:
                    if DataSourceItem.select().where(DataSourceItem.id == item_id).exists():
                        with datastore.atomic() as tx:
                            is_dataSourceItem_updated = False
                            message_sent = False
                            try:
                                q = (DataSourceItem.update({
                                                            DataSourceItem.modified: datetime.utcnow(),
                                                            DataSourceItem.title: dsi["title"],
                                                            DataSourceItem.description: dsi["description"],
                                                            DataSourceItem.iconUri: dsi["iconUri"],
                                                            DataSourceItem.type: dsi["type"],
                                                            DataSourceItem.url: dsi["url"],
                                                            DataSourceItem.updateInterval: dsi["updateInterval"],                                                            
                                                            DataSourceItem.isActive: dsi["isActive"],
                                                            DataSourceItem.isAutoCreated: False,
                                                            DataSourceItem.config: dsi["config"],
                                                            DataSourceItem.account: dsi["account"],
                                                            DataSourceItem.timezone: dsi["timezone"]
                                                            }).where(DataSourceItem.id == item_id))
                                q.execute()
                                logger.info("Updated dataSourceItem")
                                is_dataSourceItem_updated = True
                            except Exception as e:
                                logger.error("Could not update entity: {}".format(e))
                            if is_dataSourceItem_updated:
                                item = DataSourceItem.select().where(DataSourceItem.id == item_id).get()
                                dsi = DataSourceItemSchema()
                                # res = prepare_0_9(item=item)
                                producer = Producer(exchange=Config.IMPORTER_DATASOURCE_UPDATED_EXCHANGE_0_9,
                                                    type_='direct',
                                                    queue=Config.IMPORTER_DATASOURCE_UPDATED_QUEUE_0_9,
                                                    routing_key=Config.IMPORTER_DATASOURCE_UPDATED_ROUTING_KEY_0_9,
                                                    config=RabbitConfigMaker.main_rabbit_config())
                                try:
                                    EventsSender.send_0_9_event(item=item, producer=producer)
                                    message_sent = True
                                except Exception as exc:
                                    logger.error("Error sending DataSource 0.9 Updated Event: {}".format(exc))
                                    tx.rollback()
                                    logger.warning("Transaction rolled back")
                                    self.set_status(500)
                                    response = {"statusCode": 500,
                                                "description": "Could not send DataSource updated 0.9 event. Transaction "
                                                               "rolled back"}
                                    self.write(response)
                            if message_sent:
                                tx.commit()
                                self.set_status(200)
                                response = {"statusCode": 200, "description": "Success", "data": dsi.dump(item)[0]}
                                self.write(response)
                    else:
                        self.set_status(404)
                        response = {"statusCode": 404, "description": "Item not found:{}".format(item_id), "errorCode": None}
                        response = {"response": response}
                        self.write(response)
                else:
                    logger.error("Error during PUT request: {}".format(errors))
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Errors: {}".format(errors), "errorCode": None}
                    response = {"response": response}
                    self.write(response)
        else:
            self.set_status(404)
            response = {"statusCode": 404, "description": "Invalid uuid4 provided", "errorCode": None}
            response = {"response": response}
            self.write(response)

    def delete(self, item_id):
        """Delete DatasourceItem.
        ---
        description: Delete DatasourceItem.
        parameters:        
        -   name: item_id
            in: path
            description: item_id
            required: true
            type: string        
        responses:
            204:
                description: Success
        """        
        
        if DataSourceItem.select().where(DataSourceItem.id == item_id).exists():
            item = DataSourceItem.select().where(DataSourceItem.id == item_id).get()
            with datastore.atomic() as tx:
                item_deleted = False
                try:
                    item.delete_instance()
                    item_deleted = True
                except Exception as e:
                    logger.error("Cant delete dataSourceItem {}: {}".format(item.id, e))
                """
                DataSourceItem Deleted Event sending to Bus  
                """
                data = {"uid": item_id}
                producer = Producer(exchange=Config.IMPORTER_DATASOURCE_DELETED_EXCHANGE_0_9,
                                    type_='direct',
                                    queue=Config.IMPORTER_DATASOURCE_DELETED_QUEUE_0_9,
                                    routing_key=Config.IMPORTER_DATASOURCE_DELETED_ROUTING_KEY_0_9,
                                    config=RabbitConfigMaker.main_rabbit_config())
                if item_deleted:
                    message_sent = False
                    try:
                        message_sent = EventsSender.send_0_9_deleted(data=data, producer=producer)

                    except Exception as exc:
                        logger.error("Error sending DataSourceItem Updated Event: {}".format(exc))
                    """
                    End of Event
                    """
                if all([item_deleted, message_sent]):
                    self.set_status(204)
                    self.finish()
                    logger.info("Item {} deleted".format(item.id))
                else:
                    self.set_status(500)
                    response = {"statusCode": 500,
                                "description": "Error deletimg dataSourceItem {}: DataStore or RabbitMQ problems".format(item.id)}
                    self.write(response)
        else:
            self.set_status(404)
            response = {"statusCode": 404, "description": "Item not found:{}".format(item_id)}
            self.write(response)    


class DataSourceItemsSearchHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]
    FILTERS = ['status', 'type', 'result']

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self):        
        """Returns list of Datasources.
        ---
        description: Returns list of Datasources.
        parameters:
        -   name: url
            in: query
            description: url
            required: false
            type: string
        -   name: type
            in: query
            description: filters
            required: false
            type: array
        -   name: status
            in: query
            description: filters
            required: false
            type: array
        -   name: result
            in: query
            description: filters
            required: false
            type: array    
        -   name: pageSize
            in: query
            description: pageSize
            required: false
            type: integer
        -   name: page
            in: query
            description: page
            required: false
            type: integer
        -   name: sort_by
            in: query
            description: sortBy
            required: false
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                page:
                                    type: integer
                                    description: page
                                count:
                                    type: integer
                                    description: count
                                dataSources:
                                    type: integer
                                    description: dataSources list
        """
        r = re.compile(r"(desc|asc)\(([A-Za-z]+)\)", re.I)
        url = self.get_argument("url", default="")
        pageSize = self.get_argument("pageSize", default=10)
        page = self.get_argument("page", default=1)
        sortBy = self.get_argument("sortBy", default="asc(title),"
                                                       "asc(id),"
                                                       "asc(description),"
                                                       "asc(isActive),"
                                                       "asc(timezone)")
        sorts = sortBy.split(',')
        filter_status = self.get_arguments("status")
        filter_type = self.get_arguments("type")
        filter_result = self.get_arguments("result")        
        
        try:
            total_sources = DataSourceItem.select().where(DataSourceItem.url.contains(url))
        except Exception as e:
            logger.error("No dataSourceItems found: {}".format(e))
            self.set_status(404)
            logger.error("No dataSourceItems found: {}".format(e))
            response = {"statusCode": 404, "description": "No dataSources found: {}".format(e)}
            self.write(response)
            return
        if total_sources:
            if sortBy:
                default_sorted_sources = total_sources.order_by(DataSourceItem.title.asc(),
                                                            DataSourceItem.id.asc(),
                                                            DataSourceItem.description.asc(),
                                                            DataSourceItem.isActive.asc(),
                                                            DataSourceItem.timezone.asc(),
                                                        )

                """Sort DataSources"""
                for sort in sorts:
                    f = re.match(r, sort)
                    sort_direction = f.group(1)
                    filter = f.group(2)
                    if filter not in DataSourceItem._meta.fields:
                        logger.warning("Field {} is forbidden".format(filter))
                    else:
                        field = getattr(DataSourceItem, filter)
                        if sort_direction != 'asc':
                            logger.info("PageSize: {}, Page: {}, Sort direction: {}, Field: {}".format(pageSize,
                                                                                                       page,
                                                                                                       sort_direction,
                                                                                                       field))
                            field = field.desc()
                            default_sorted_sources = default_sorted_sources.order_by(field)
            else:
                default_sorted_sources = total_sources

            sources_list = []
            dsis_schema = DataSourceItemSearchSchema()
            for source in default_sorted_sources:
                d = {}                
                d = dsis_schema.dump(source)[0]
                d.executionStatus = ""
                d.lastResult = ""
                
                if not source.isActive and source.updateInterval > 0:
                     d.executionStatus = "stopped"
                if source.updateInterval == 0:
                     d.executionStatus = "notconfigure"
                if source.isActive and Job.select().where(Job.dataSourceItemUid == source.id).exists():
                    lastJobs = Job.select().where(Job.dataSourceItemUid == source.id).order_by(Job.created.desc())
                    d.executionStatus = "planned"
                    if lastJobs[0].done_by_worker:
                        d.lastResult = "success" if lastJobs[0].success else "error"
                    elif lastJobs[0].read_by_worker:
                            d.executionStatus = "inprogress"
                            if lastJobs.count() > 1:
                                d.lastResult = "success" if lastJobs[1].success else "error"
                    now = datetime.utcnow()
                    time_passed = now - lastJobs[0].created
                    seconds_passed = time_passed.total_seconds()
                    d.nextLaunch = source.updateInterval - seconds_passed
                sources_list.append(dsis_schema.dump(d)[0])

            logger.warning("total_sources: {}".format(total_sources))
            logger.warning("sources_list: {}".format(sources_list))

            filtered_sources_list = [item for item in sources_list if
                ((len(filter_status) > 0 and item["executionStatus"] in filter_status) or (len(filter_status) == 0)) and
                ((len(filter_type) > 0 and item["type"] in filter_type) or (len(filter_type) == 0)) and
                ((len(filter_result) > 0 and item["lastResult"] in filter_result) or (len(filter_result) == 0))]
                                       
            total_sources_count = len(filtered_sources_list)

            if total_sources_count == 0:
                logger.error("No dataSourceItems found.")
                self.set_status(404)
                response = {"statusCode": 404, "description": "No dataSources found."}
                self.write(response)
                return
            
            """Paginate results"""
            paginated_datasources = [filtered_sources_list[i:i+pageSize] for i in range(0, len(filtered_sources_list), int(pageSize))]
            #sources_list.paginate(int(page), int(pageSize))

            logger.warning("paginated_datasources: {}".format(paginated_datasources))
           

            response = {"statusCode": 200, "description": "Success", "data": {
                                                                        "pageSize": int(pageSize),
                                                                        "page": int(page),
                                                                        "count": total_sources_count,
                                                                        "dataSourceItems": paginated_datasources[int(page)-1]}}
            self.write(response)


class ItemsHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self):
        """Returns info about DataSourceItems.
        ---
        description: Returns list of DataSourceItems. If "all" parameter passed, all(incl. isActive=false) is returned. Used by Scheduler
        parameters:
        -   name: all
            in: path
            description: all
            required: true
            type: string
        -   name: social
            in: path
            description: social
            required: true
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        count:
                            type: integer
                            description: count
                        items:
                            type: array
                            description: items list                      
                        errorCode:
                            type: string
                            description: errorCode
        """
        social_types = Config.SOCIAL_TYPES
        all_items = None
        get_all = self.get_query_argument("all", "0")
        get_social = self.get_query_argument("social", "0")

        if get_all == "1":
            if DataSourceItem.select().exists():
                all_items = DataSourceItem.select()
        else:
            if DataSourceItem.select().where(DataSourceItem.isActive==True).exists():
                if get_social == "1":
                    all_items = DataSourceItem.select().where(DataSourceItem.type.in_(social_types),
                                                              DataSourceItem.isActive==True)
                elif get_social == "0":
                    all_items = DataSourceItem.select().where(DataSourceItem.isActive==True,
                                                              # DataSourceItem.type.not_in(social_types)
                                                              )
        items_list = []
        if all_items:
            now = datetime.utcnow()
            for item in all_items:
                time_passed = now - item.updated
                seconds_passed = time_passed.total_seconds()
                if seconds_passed >= item.updateInterval:
                    d = {}
                    DSItemSchema = DataSourceItemSchema()
                    d["dataSourceItem"] = DSItemSchema.dump(item)[0]
                    items_list.append(d)
                    q = DataSourceItem.update(
                        {DataSourceItem.updated: datetime.utcnow()}).where(DataSourceItem.id == item.id)
                    q.execute()

            self.set_status(200)
            response = {"statusCode": 200, "count": len(items_list), "items": items_list, "errorCode": None}
            response = {"response": response}
            self.write(response)
        else:
            self.set_status(200)
            response = {"statusCode": 200, "count": 0, "items": [], "errorCode": None}
            response = {"response": response}
            self.write(response)


class FileRequestHandler(BaseHandler):

    def prepare(self):
        if filestore.is_closed():
            filestore.connect()
        return super(FileRequestHandler, self).prepare()

    def on_finish(self):
        if not filestore.is_closed():
            filestore.close()
        return super(FileRequestHandler, self).on_finish()


class DataSourceFullItemsHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self):        
        """Returns list of Datasources with full items.
        ---
        description: Returns list of Datasources with full items.
        parameters:
        -   name: pageSize
            in: query
            description: pageSize
            required: false
            type: integer
        -   name: page
            in: query
            description: page
            required: false
            type: integer
        -   name: sort_by
            in: query
            description: sort_by
            required: false
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: string
                            description: description                      
                        data:
                            type: object
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                page:
                                    type: integer
                                    description: page
                                count:
                                    type: integer
                                    description: count
                                dataSources:
                                    type: array
                                    description: dataSources
        """
        r = re.compile(r"(desc|asc)\(([A-Za-z]+)\)", re.I)
        ds_schema = DataSourceSchema()
        pageSize = self.get_argument("pageSize", default=10)
        page = self.get_argument("page", default=1)
        sort_by = self.get_argument("sort_by", default="asc(title),"
                                                       "asc(id),"
                                                       "asc(description),"
                                                       "asc(isActive),"
                                                       "asc(isQuasy),"
                                                       "asc(timezone)")
        sorts = sort_by.split(',')
        try:
            total_sources = DataSource.select()
        except Exception as e:
            logger.error("No dataSources found: {}".format(e))
            self.set_status(404)
            logger.error("No dataSources found: {}".format(e))
            response = {"statusCode": 404, "description": "No dataSources found: {}".format(e)}
            self.write(response)
            return
        if total_sources:
            total_sources_count = total_sources.count()

            default_sorted_sources = total_sources.order_by(DataSource.title.asc(),
                                                        DataSource.id.asc(),
                                                        DataSource.description.asc(),
                                                        DataSource.isActive.asc(),
                                                        DataSource.isQuasy.asc(),
                                                        DataSource.timezone.asc(),
                                                        )
            """Sort DataSources"""
            for sort in sorts:
                f = re.match(r, sort)
                sort_direction = f.group(1)
                filter = f.group(2)
                if filter not in DataSource._meta.fields:
                    logger.warning("Field {} is forbidden".format(filter))
                else:
                    field = getattr(DataSource, filter)
                    if sort_direction != 'asc':
                        logger.info("PageSize: {}, Page: {}, Sort direction: {}, Field: {}".format(pageSize,
                                                                                                   page,
                                                                                                   sort_direction,
                                                                                                   field))
                        field = field.desc()
                        default_sorted_sources = default_sorted_sources.order_by(field)
            """Paginate results"""
            paginated_datasources = default_sorted_sources.paginate(int(page), int(pageSize))
            sources_list = []
            dsi_schema = DataSourceItemSchema()
            for source in paginated_datasources:
                d = {}
                items_list = []
                items = DataSourceItem.select().where(DataSourceItem.dataSource == source.id)
                for item in items:
                    items_list.append(dsi_schema.dump(item)[0])
                d = ds_schema.dump(source)[0]
                d["dataSourceItems"] = items_list
                sources_list.append(d)
            response = {"statusCode": 200, "description": "Success", "data": {
                                                                        "pageSize": int(pageSize),
                                                                        "page": int(page),
                                                                        "count": total_sources_count,
                                                                        "dataSources": sources_list}}
            self.write(response)


class DataSourcesHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "POST", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self):        
        """Returns list of Datasources.
        ---
        description: Returns list of Datasources.
        parameters:
        -   name: pageSize
            in: path
            description: pageSize
            required: false
            type: integer
        -   name: page
            in: path
            description: page
            required: false
            type: integer
        -   name: sort_by
            in: path
            description: sort_by
            required: false
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: data
                            properties:
                                pageSize:
                                    type: integer
                                    description: pageSize
                                page:
                                    type: integer
                                    description: page
                                count:
                                    type: integer
                                    description: count
                                dataSources:
                                    type: integer
                                    description: dataSources list
        """
        r = re.compile(r"(desc|asc)\(([A-Za-z]+)\)", re.I)
        ds_schema = DataSourceSchema()
        pageSize = self.get_argument("pageSize", default=10)
        page = self.get_argument("page", default=1)
        sort_by = self.get_argument("sort_by", default="asc(title),"
                                                       "asc(id),"
                                                       "asc(description),"
                                                       "asc(isActive),"
                                                       "asc(isQuasy),"
                                                       "asc(timezone)")
        sorts = sort_by.split(',')
        try:
            total_sources = DataSource.select()
        except Exception as e:
            logger.error("No dataSources found: {}".format(e))
            self.set_status(404)
            logger.error("No dataSources found: {}".format(e))
            response = {"statusCode": 404, "description": "No dataSources found: {}".format(e)}
            self.write(response)
            return
        if total_sources:

            total_sources_count = total_sources.count()
            if sort_by:
                default_sorted_sources = total_sources.order_by(DataSource.title.asc(),
                                                            DataSource.id.asc(),
                                                            DataSource.description.asc(),
                                                            DataSource.isActive.asc(),
                                                            DataSource.isQuasy.asc(),
                                                            DataSource.timezone.asc(),
                                                        )

                """Sort DataSources"""
                for sort in sorts:
                    f = re.match(r, sort)
                    sort_direction = f.group(1)
                    filter = f.group(2)
                    if filter not in DataSource._meta.fields:
                        logger.warning("Field {} is forbidden".format(filter))
                    else:
                        field = getattr(DataSource, filter)
                        if sort_direction != 'asc':
                            logger.info("PageSize: {}, Page: {}, Sort direction: {}, Field: {}".format(pageSize,
                                                                                                       page,
                                                                                                       sort_direction,
                                                                                                       field))
                            field = field.desc()
                            default_sorted_sources = default_sorted_sources.order_by(field)
            else:
                default_sorted_sources = total_sources
            """Paginate results"""
            paginated_datasources = default_sorted_sources.paginate(int(page), int(pageSize))
            sources_list = []
            for source in paginated_datasources:
                d = {}
                items_list = []
                items = DataSourceItem.select().where(DataSourceItem.dataSource == source.id)
                for item in items:
                    items_list.append(str(item.id))
                d = ds_schema.dump(source)[0]
                d["dataSourceItems"] = items_list
                sources_list.append(d)
            response = {"statusCode": 200, "description": "Success", "data": {
                                                                        "pageSize": int(pageSize),
                                                                        "page": int(page),
                                                                        "count": total_sources_count,
                                                                        "dataSources": sources_list}}
            self.write(response)

    def post(self):        
        """Creates a new DataSource with Items.
        ---
        description: Creates a new DataSource with Items.
        parameters:
        -   name: dataSource
            in: body
            description: dataSource
            required: false
            type: DataSourceSchema
        -   name: dataSourceItems
            in: body
            description: dataSourceItems
            required: false
            type: DataSourceItemSchema       
        responses:
            201:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode                       
                        description:
                            type: string
                            description: description                      
                        data:
                            type: array
                            description: dataSources list                            
        """
        raw_body = None
        try:
            raw_body = json_decode(self.request.body)
        except Exception as e:
            logger.error(("Error reading POST body: {}".format(e)))
            self.set_status(400)
            logger.error("Error in request body: {}".format(e))
            response = {"statusCode": 400, "description": "No dataSources found: {}".format(e)}
            self.write(response)
            return
        if raw_body:
            logger.debug("RAW body: {}".format(raw_body))
            post_data_schema = DataSourcePostSchema()
            body, errors = post_data_schema.load(raw_body)
            logger.info("body: {}".format(body))
            if len(errors) == 0:
                items_created = 0
                messages_sent = 0
                items_count = 0
                new_dsi = None
                ds_dump = None
                with datastore.atomic() as tx:
                    items_list = []
                    datasource = body["dataSource"]
                    if "dataSourceItems" in body:
                        datasourceitems = body["dataSourceItems"]
                    else:
                        datasourceitems = None
                    fan_dataSourceItems = []
                    if not DataSource.select().where(DataSource.id == datasource["id"]).exists():
                        new_ds = DataSource(**datasource)
                        # new_ds.id = uuid4()
                        new_ds.save(force_insert=True)
                        if new_ds:
                            items_count = len(datasourceitems)
                            for item in datasourceitems:
                                if not all([DataSourceItem.select().where(DataSourceItem.id == item["id"]).exists(),
                                            DataSourceItem.select().where(DataSourceItem.url == item["url"]).exists()]):
                                    item_created = False
                                    if "updateInterval" not in item:
                                        item["updateInterval"] = new_ds.updateInterval
                                    try:
                                        new_dsi = DataSourceItem(**item)
                                        new_dsi.dataSource = new_ds.id
                                        # new_dsi.isAutoCreated = False
                                        # new_dsi.id = uuid4()
                                        new_dsi.updated = datetime.utcnow()
                                        new_dsi.save(force_insert=True)
                                        item_created = True
                                        items_created += 1
                                    except Exception as e:
                                        logger.error("Could not save DataSourceItem {}".format(e))
                                    if item_created:
                                        """send 0.9 event"""
                                        dsi_schema = DataSourceItemSchema()
                                        dsi_result = dsi_schema.dump(new_dsi)[0]
                                        items_list.append(dsi_result)
                                        fan_dataSourceItems.append(dsi_result)
                                        # ds_msg = prepare_0_9(new_dsi)
                                        producer = Producer(exchange=Config.IMPORTER_DATASOURCE_CREATED_EXCHANGE_0_9,
                                                            type_='direct',
                                                            queue=Config.IMPORTER_DATASOURCE_CREATED_QUEUE_0_9,
                                                            routing_key=Config.IMPORTER_DATASOURCE_CREATED_ROUTING_KEY_0_9,
                                                            config=RabbitConfigMaker.main_rabbit_config())
                                        try:
                                            EventsSender.send_0_9_event(item=new_dsi, producer=producer)
                                            messages_sent += 1
                                            logger.info("Sent 0.9 event: {}".format(EventsSender.prepare_0_9(new_dsi)))
                                        except Exception as exc:
                                            logger.error("Error sending DataSource Created Event: {}".format(exc))
                                        """
                                        End of v0.9
                                        """
                                        """v1.0"""
                                        """End of v1.0"""
                                else:
                                    logger.error("DataSourceItemUid {} already exists".format(item["id"]))
                                    tx.rollback()
                                    self.set_status(400)
                                    response = {"statusCode": 400,
                                                "description": "DataSourceItemUid {} already exists".format(
                                                    item["id"])}
                                    self.write(response)
                                    return
                            ds_schema = DataSourceSchema()
                            ds_result = ds_schema.dump(new_ds)[0]
                            ds_dump = ds_result
                            """
                            DataSource created Event sending to Bus v1.0
                            """
                            """
                            End of Event v1.0
                            """
                        if all([items_created == items_count, messages_sent == items_count]):
                            tx.commit()
                            ds_dump["dataSourceItems"] = items_list
                            response = {"statusCode": 201, "description": "Created", "data": {"dataSource": ds_dump}}
                            self.set_status(201)
                            self.set_header('Location', "http://" + Config.IMPORTER_API_HOST + ":" + str(Config.IMPORTER_API_PORT) +
                                            self.reverse_url("api:datasource", new_ds.id))
                            self.write(response)
                        else:
                            tx.rollback()
                            logger.error("Transaction rolled back")
                            self.set_status(500)
                            response = {"statusCode": 500, "description": "Could not create dataSource and/or send 0.9 event"}
                            self.write(response)
                    else:
                        tx.rollback()
                        logger.error("DataSourceUid {} already exists".format(datasource["id"]))
                        self.set_status(400)
                        response = {"statusCode": 400,
                                    "description": "DataSourceUid {} already exists".format(datasource["id"])}
                        self.write(response)

            else:
                logger.error("Error validating request body: {}".format(errors))
                self.set_status(400)
                response = {"statusCode": 400, "description": "Error in POST request: {}.".format(errors)}
                self.write(response)


class DataSourceHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "PUT", "PATCH", "DELETE", "OPTIONS"]

    def options(self, datasource_id):
        self.set_header('Access-Control-Allow-Methods', 'GET, PUT, PATCH, DELETE, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self, datasource_id):        
        """Returns info about Datasource.
        ---
        description: Returns info about Datasource.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: string
                            description: description                      
                        data:
                            type: string
                            description: dataSourceItems
        """
        source = None
        dataSourceSchema = DataSourceSchema()
        dataSourceItemSchema = DataSourceItemSchema()
        if is_uuid4(datasource_id):
            if DataSource.select().where(DataSource.id == datasource_id).exists():
                try:
                    source = DataSource.select().where(DataSource.id == uuid.UUID(datasource_id)).get()
                except peewee.DoesNotExist as e:
                    logger.error("Error getting DataSource from DB:{}".format(e))
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Error getting item from DB:{}".format(e)}
                    self.write(response)
                if source != None:
                    dumped_source = dataSourceSchema.dump(source)[0]
                    datasourceitems = DataSourceItem.select().where(DataSourceItem.dataSource == source)
                    data = dumped_source
                    items_list = []
                    for datasourceitem in datasourceitems:
                        items_list.append(dataSourceItemSchema.dump(datasourceitem)[0])
                    data["dataSourceItems"] = items_list
                    self.set_status(200)
                    response = {"statusCode": 200,
                                "description": "Success",
                                "data": data}
                    self.write(response)
                else:
                    self.set_status(404)
                    response = {"statusCode": 404, "description": "Item {} not found".format(datasource_id)}
                    self.write(response)

            else:
                self.set_status(404)
                response = {"statusCode": 404,
                            "description": "DataSource does not exist:{}".format(datasource_id)}
                self.write(response)
        else:
            self.set_status(404)
            response = {"statusCode": 404,
                        "description": "DataSource not found. Invalid uuid4 provided: {}".format(datasource_id),}
            self.write(response)

    def patch(self, datasource_id):        
        """Patch datasource.
        ---
        description: Patch datasource.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: string
                            description: description                      
                        data:
                            type: string
                            description: dataSourceItems
        """
        if is_uuid4(datasource_id):
            allowed_fields = ["isQuasy", "iconUri", "title", "description", "updateInterval", "isActive", "timezone"]
            raw_body = None
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in request body: {}".format(e))
                self.set_status(400)
                response = {"statusCode": 400, "description": "Errors in request body: {}".format(e)}
                self.write(response)
                return
            if raw_body:
                ds_patch_schema = DataSourcePatchSchema()
                body, errors = ds_patch_schema.load(raw_body)
                error_list = []
                for key in list(body.keys()):
                    if key in allowed_fields:
                        pass
                    else:
                        del body[key] 
                        logger.error("Field {} not found or not allowed".format(key)) #TODO: fix incoming fields or store them
                        #d = {key: value}
                        #error_list.append(d)

                if all([len(error_list) == 0, len(errors) == 0]):
                    logger.debug("No errors in request body")
                    logger.debug("body: {}".format(body))
                    try:
                        ds = DataSource.select().where(DataSource.id == datasource_id).get()
                    except peewee.DoesNotExist as e:
                        logger.error("Couldn't find DataSource: {}".format(e))
                        self.set_status(200)
                        response = {"statusCode": 200, "description": " DataSource not found. {}".format(e)}
                        self.write(response)
                        return
                    q = DataSource.update(**body).where(DataSource.id == datasource_id)
                    q.execute()
                    q = DataSource.update({DataSource.modified: datetime.utcnow()}).where(DataSource.id == datasource_id)
                    q.execute()
                    logger.info("DataSource updated")
                    ds_schema = DataSourceSchema()
                    updated_ds = DataSource.select().where(DataSource.id == datasource_id).get()
                    """update all dataSourceItems isActive field"""
                    dsitems = DataSourceItem.select().where(DataSourceItem.dataSource==datasource_id)
                    logger.info("DataSource has {} items".format(len(dsitems)))
                    q = DataSourceItem.update({DataSourceItem.isActive: updated_ds.isActive}).where(
                                               DataSourceItem.dataSource==datasource_id
                                               )
                    q.execute()
                    logger.debug("isActive status updated for {} items and is {}".format(
                                                                                len(dsitems),
                                                                                updated_ds.isActive
                                                                            ))
                    self.set_status(200)
                    response = {"statusCode": 200, "description": "Success", "data": ds_schema.dump(updated_ds)[0]}
                    self.write(response)
                else:
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Bad request. {}, {}".format(errors, error_list)}
                    self.write(response)
        else:
            self.set_status(400)
            response = {"statusCode": 400, "description": "Bad request: {}".format(datasource_id)}
            self.write(response)

    def put(self, datasource_id):
        """UPDATE Datasource's data only.
        ---
        description: UPDATE Datasource's data only.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        -   name: title
            in: body
            description: title
            required: true
            type: string
        -   name: description
            in: body
            description: description
            required: true
            type: string
        -   name: updateInterval
            in: body
            description: updateInterval
            required: true
            type: integer
        -   name: isActive
            in: body
            description: isActive
            required: true
            type: boolean
        -   name: isQuasy
            in: body
            description: isQuasy
            required: true
            type: boolean
        -   name: iconUri
            in: body
            description: iconUri
            required: true
            type: string
        -   name: timezone
            in: body
            description: timezone
            required: false
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: string
                            description: description                      
                        data:
                            type: string
                            description: dataSourceItems
        """

        if is_uuid4(datasource_id):
            raw_body = None
            message_sent = False
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in request body: {}".format(e))
                self.set_status(400)
                response = {"statusCode": 400, "description": "Errors in request body: {}".format(e)}
                self.write(response)
                return
            if raw_body:
                ds_schema = DataSourcePutSchema()
                dsi_schema = DataSourceItemSchema()
                datasource, errors = ds_schema.load(raw_body)
                if len(errors) == 0:
                    if DataSource.select().where(DataSource.id == datasource_id).exists():
                        with datastore.atomic() as tx:
                            source_updated = False
                            try:
                                q = (DataSource.update({DataSource.title: datasource["title"],
                                                        DataSource.description: datasource["description"],
                                                        DataSource.iconUri: datasource["iconUri"],
                                                        DataSource.isQuasy: datasource["isQuasy"],
                                                        DataSource.isActive: datasource["isActive"],
                                                        DataSource.timezone: datasource["timezone"],
                                                        DataSource.modified: datetime.utcnow(),
                                                        DataSource.updateInterval: datasource["updateInterval"]}).where(DataSource.id == datasource_id))
                                q.execute()
                                source_updated = True
                                logger.info("DataSource updated")
                            except Exception as e:
                                logger.error("Error updating DataSource: {}".format(e))
                            if source_updated:
                                message_sent = False
                                updated_ds = DataSource.select().where(DataSource.id == datasource_id).get()
                                """update all dataSourceItems isActive field"""
                                dsitems = DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id)
                                q = DataSourceItem.update({DataSourceItem.isActive: updated_ds.isActive}).where(
                                    DataSourceItem.dataSource==datasource_id)
                                q.execute()
                                logger.info("isActive status updated for {} items and is {}".format(len(dsitems), datasource["isActive"]))
                                ds = DataSourceSchema()
                                result, error = ds.dump(updated_ds)
                                """ v1.0
                                DataSource Updated Event sending to Bus
                                """
                                message_sent = True
                                """
                                End of Event
                                """
                            if message_sent:
                                tx.commit()
                                items = DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id)
                                items_list = []
                                for item in items:
                                    items_list.append(dsi_schema.dump(item)[0])
                                result["dataSourceItems"] = items_list
                                self.set_status(200)
                                response = {"statusCode": 200, "description": "Success", "data":result}
                                self.write(response)
                    else:
                        logger.error("DataSource {} not found: ".format(datasource_id))
                        self.set_status(404)
                        response = {"statusCode": 404, "description": "DataSource {} not found: ".format(datasource_id)}
                        self.write(response)
                else:
                    logger.error("Errors in request: {}".format(errors))
                    self.set_status(400)
                    response = {"statusCode": 400, "description": "Errors in request: {}".format(errors)}
                    self.write(response)
        else:
            logger.error("DataSource {} not found: ".format(datasource_id))
            self.set_status(404)
            response = {"statusCode": 404,
                        "description": "DataSource {} not found. Invalid uid provided.".format(datasource_id)}
            self.write(response)

    def delete(self, datasource_id):
        """DELETE Datasource.If linked DataSourceItems exist, they will be deleted.
        ---
        description: DELETE Datasource.If linked DataSourceItems exist, they will be deleted.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        responses:
            204:
                description: Success                
        """
        if is_uuid4(datasource_id):
            if DataSource.select().where(DataSource.id == datasource_id).exists():
                with datastore.atomic() as tx:
                    datasource = DataSource.select().where(DataSource.id == datasource_id).get()
                    if DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id).exists():
                        items = DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id)
                        items_deleted = 0
                        items_deleted = False
                        messages_sent = 0
                        for item in items:
                            item.delete_instance()
                            items_deleted += 1
                            item_deleted = True
                            logger.info("Item {} deleted".format(item.id))
                            if item_deleted:
                                """v.0.9
                                DataSource deleted Event sending to Bus 
                                """
                                ds_dump = {"uid": str(item.id)}
                                producer = Producer(exchange=Config.IMPORTER_DATASOURCE_DELETED_EXCHANGE_0_9,
                                                    type_='direct',
                                                    queue=Config.IMPORTER_DATASOURCE_DELETED_QUEUE_0_9,
                                                    routing_key=Config.IMPORTER_DATASOURCE_DELETED_ROUTING_KEY_0_9,
                                                    config=RabbitConfigMaker.main_rabbit_config()
                                                    )
                                try:
                                    EventsSender.send_0_9_deleted(data=ds_dump, producer=producer)
                                    messages_sent += 1
                                except Exception as exc:
                                    logger.error("Error sending DataSource Deleted Event: {}".format(exc))
                                """
                                End of Event
                                """
                        source_deleted = False
                        try:
                            datasource.delete_instance()
                            logger.info("DataSource {} deleted.".format(datasource_id))
                            source_deleted = True
                        except Exception as e:
                            logger.error("Could not delete dataSource: {}".format(e))

                        """v1.0
                        DataSource deleted Event sending to Bus
                        """
                        """
                        End of Event
                        """
                        if all([source_deleted == True, messages_sent == items_deleted]):
                            self.set_status(204)
                            tx.commit()
                        else:
                            tx.rollback()
                            logger.error("Transaction rolled back")
                            self.set_status(500)
                            response = {"statusCode": 500,
                                        "description": "Could not send DataSource Deleted 0.9 event. Transaction rolled back"}
                            self.write(response)
            else:
                self.set_status(404)
                response = {"statusCode": 404, "description": "DataSource not found:{}".format(datasource_id)}
                self.write(response)
        else:
            self.set_status(404)
            response = {"statusCode": 404,
                        "description": "DataSource not found. Invalid uuid4 provided: {}".format(datasource_id)}
            self.write(response)


class DataSourcesItemsHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "POST", "OPTIONS"]

    def options(self, datasource_id):
        self.set_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self, datasource_id):        
        """Returns list of DataSource's items.
        ---
        description: Returns list of DataSource's items.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        count:
                            type: integer
                            description: count
                        description:
                            type: string
                            description: description                      
                        items:
                            type: array
                            description: items_list
                        errorCode:
                            type: string
                            description: errorCode
        """
        if is_uuid4(datasource_id):
            if DataSource.select().where(DataSource.id == datasource_id).exists():
                datasource = DataSource.select().where(DataSource.id == datasource_id).get()
                if DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id).exists():
                    items = DataSourceItem.select().where(DataSourceItem.dataSource == datasource_id)
                # "query exists"
                    if items:
                        items_list = []
                        for item in items:
                            itemSchema = DataSourceItemSchema()
                            result, errors = itemSchema.dump(item)
                            items_list.append(result)
                        self.set_status(200)
                        response = {"statusCode": 200, "count": len(items_list), "items": items_list, "errorCode": None}
                        response = {"response": response}
                        self.write(response)
                    else:
                        self.set_status(404)
                        response = {"statusCode": 404,
                                    "description": "DataSourceItem(s) for DataSource {} not found".format(datasource_id),
                                    "errorCode": None}
                        response = {"response": response}
                        self.write(response)
                else:
                    self.set_status(400)
                    response = {"statusCode": 400,
                                "description": "DataSourceItems for DataSource {} do not exist".format(datasource_id),
                                "errorCode": None}
                    response = {"response": response}
                    self.write(response)
            else:
                self.set_status(404)
                response = {"statusCode": 404, "description": "DataSource does not exist:{}".format(datasource_id), "errorCode": None}
                response = {"response": response}
                self.write(response)
        else:
            self.set_status(404)
            response = {"statusCode": 404, "description": "DataSource does not exist. Invalid uuid4 provided: {}".format(datasource_id),
                        "errorCode": None}
            response = {"response": response}
            self.write(response)

    def post(self, datasource_id):        
        """Creates a new DataSourceItem.
        ---
        description: Creates a new DataSourceItem.
        parameters:
        -   name: datasource_id
            in: path
            description: datasource_id
            required: true
            type: string
        responses:
            201:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        data:
                            type: string
                            description: result
                        description:
                            type: string
                            description: description
        """
        raw_body = None
        if is_uuid4(datasource_id):
            try:
                raw_body = json_decode(self.request.body)
            except Exception as e:
                logger.error("Error in POST body: {}".format(e))
                self.set_status(400)
                logger.error("Error in request body: {}".format(e))
                response = {"statusCode": 400, "description": "Error in request body: {}".format(e),
                            "errorCode": None}
                response = {"response": response}
                self.write(response)
                return
            if raw_body:
                message_sent = False
                logger.info("raw body: {}".format(raw_body))
                dsi_schema = DataSourceItemSchema()
                datasourceitem, errors = dsi_schema.load(raw_body)
                if len(errors) == 0:
                    if DataSource.select().where(DataSource.id == datasource_id).exists():
                        if not DataSourceItem.select().where(DataSourceItem.id == datasourceitem["id"]).exists():
                            with datastore.atomic() as tx:
                                item_saved = False
                                try:
                                    new_dataSourceItem = DataSourceItem(**datasourceitem)
                                    new_dataSourceItem.updated = datetime.utcnow()
                                    new_dataSourceItem.dataSource = datasource_id
                                    # new_dataSourceItem.id = uuid4()  # must be generated by UI
                                    new_dataSourceItem.save(force_insert=True)
                                    logger.info("Item saved")
                                    item_saved = True
                                except Exception as e:
                                    logger.error(("Error saving new dataSourceItem: {}".format(e)))
                                logger.info("new_dataSourceItem: {}".format(new_dataSourceItem))
                                # new_item = dsi_schema.load(new_dataSourceItem)[0]
                                result, error = dsi_schema.dump(new_dataSourceItem)
                                """v.0.9
                                DataSource created Event sending to Bus 
                                """
                                # ds_msg = prepare_0_9(item=new_dataSourceItem)
                                if item_saved:
                                    producer = Producer(exchange=Config.IMPORTER_DATASOURCE_CREATED_EXCHANGE_0_9,
                                                        type_='direct',
                                                        queue=Config.IMPORTER_DATASOURCE_CREATED_QUEUE_0_9,
                                                        routing_key=Config.IMPORTER_DATASOURCE_CREATED_ROUTING_KEY_0_9,
                                                        config=RabbitConfigMaker.main_rabbit_config()
                                                        )
                                    try:
                                        EventsSender.send_0_9_event(item=new_dataSourceItem, producer=producer)
                                        message_sent = True
                                    except Exception as exc:
                                        logger.error("Error sending DataSource Created Event: {}".format(exc))
                                """
                                End of Event
                                """
                                if message_sent:
                                    tx.commit()
                                    response = {"statusCode": 201,
                                                "description": "Success",
                                                "data": result}
                                    self.set_status(201)
                                    self.set_header('Location', "http://" + Config.IMPORTER_API_HOST + ":" + str(Config.IMPORTER_API_PORT) +
                                                    self.reverse_url("api:datasourceitem", datasource_id, new_dataSourceItem.id))
                                    self.write(response)
                                else:
                                    tx.rollback()
                                    logger.info("Transaction rolled back")
                                    self.set_status(500)
                                    response = {"statusCode": 500,
                                                "description": "Could not send DataSource Created 0.9 event. Transaction "
                                                               "rolled back"}
                                    self.write(response)
                        else:
                            self.set_status(400)
                            logger.error("DataSourceItemUid {} already exists".format(datasourceitem["id"]))
                            response = {"statusCode": 400,
                                        "description": "DataSourceItemUid {} already exists".format(datasourceitem["id"])}
                            self.write(response)
                    else:
                        self.set_status(404)
                        logger.error("DataSource {} does not exist".format(datasource_id))
                        response = {"statusCode": 404, "description": "DataSource {} does not exist".format(datasource_id)}
                        self.write(response)
                else:
                    self.set_status(400)
                    logger.error("Error creating DataSourceItem: {}".format(errors))
                    response = {"statusCode": 400, "description": "Error creating DataSourceItem: {}".format(errors)}
                    self.write(response)
        else:
            self.set_status(404)
            logger.error("DataSource does not exist")
            response = {"statusCode": 404, "description": "Invalid uuid4 provided: {}".format(datasource_id)}
            self.write(response)


class DataSourcesGetByUidsHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get_uids(self, uids):
        uids = uids.replace("[", "")
        uids = uids.replace("]", "")
        uids_list = uids.split(",")
        return uids_list

    def get(self):
        """Get DataSources By Uids.
        ---
        description: Get DataSources By Uids.
        parameters:
        -   name: uids
            in: body
            description: uids
            required: true
            type: array
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: integer
                            description: description
                        data:
                            type: string
                            description: dataSourceItems
        """
        ds_schema = DataSourceSchema()
        dsi_schema = DataSourceItemSchema()
        uids = self.get_argument("uids", default=[])
        uids_list = self.get_uids(uids)
        data = []
        if len(uids_list) > 0:
            logger.info("uids provided: {}".format(uids_list))
            for uid in uids_list:
                if is_uuid4(uid):
                    logger.debug("{} is uuid".format(uid))
                    if DataSource.select().where(DataSource.id == uid).exists():
                        dataSource = DataSource.select().where(DataSource.id == uid).get()
                        dumped_source = ds_schema.dump(dataSource)[0]
                        dataSourceItems = DataSourceItem.select().where(DataSourceItem.dataSource == uid)
                        items_list = []
                        for item in dataSourceItems:
                            dumpled_item = dsi_schema.dump(item)[0]
                            items_list.append(dumpled_item)
                        dumped_source["dataSourceItems"] = items_list
                        data.append(dumped_source)
        self.set_status(200)
        response = {"statusCode": 200, "description": "Success", "data": data}
        self.write(response)


class DataSourcesItemsGetByUidsHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get_uids(self, uids):
        uids = uids.replace("[", "")
        uids = uids.replace("]", "")
        uids_list = uids.split(",")
        return uids_list

    def get(self):
        """Get DataSources By Uids.
        ---
        description: Get DataSources By Uids.
        parameters:
        -   name: uids
            in: body
            description: uids
            required: true
            type: array
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: integer
                            description: description
                        data:
                            type: string
                            description: dataSourceItems
        """
        dsi_schema = DataSourcesItemsByUidsSchema()
        uids = self.get_argument("uids", default="[]")
        uids_list = self.get_uids(uids)
        items_list = []
        if len(uids_list) > 0:
            logger.info("uids provided: {}".format(uids_list))
            for uid in uids_list:
                if is_uuid4(uid):
                    logger.debug("{} is uuid".format(uid))
                    if DataSourceItem.select().where(DataSourceItem.id == uid).exists():
                        DSItem = DataSourceItem.select().where(DataSourceItem.id == uid).get()
                        dumped_item = dsi_schema.dump(DSItem)[0]
                        items_list.append(dumped_item)
            self.set_status(200)
            response = {"statusCode": 200, "description": "Success", "data": items_list}
            self.write(response)


class DataSourcesItemsFilterHandler(DataRequestHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]
    FILTERS = ['url', 'type']

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get_filters(self, f):
        pass

    def get(self):
        """Get DataSourcesItems By Filters.
        ---
        description: Get DataSources By Uids.
        parameters:
        -   name: filters
            in: path
            description: filters
            required: true
            type: array
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:                       
                        data:
                            type: string
                            description: dataSourceItems
        """
        filters = self.get_argument("filters", default="{}")
        logger.debug("filters: {}".format(filters))
        try:
            filters = json.loads(filters)
        except Exception as e:
            logger.warning("Incorrect filters param: {}".format(e))
            self.set_status(400)
            self.write("Incorrect filters param: {}".format(e))
            return
        items_list = []
        items = None
        if all(["url" in filters, "type" in filters]):
            items = DataSourceItem.select().where(DataSourceItem.url.contains(filters["url"]) &
                                                  DataSourceItem.type.contains(filters["type"]))
        elif all(["url" in filters, "type" not in filters]):
            items = DataSourceItem.select().where(DataSourceItem.url.contains(filters["url"]))
        elif all(["url" not in filters, "type" in filters]):
            items = DataSourceItem.select().where(DataSourceItem.type.contains(filters["type"]))
        elif all(["url" not in filters, "type" not in filters]):
            items = DataSourceItem.select()
        logger.debug("Found {}".format(len(items)))
        if items:
            dsi_schema = DataSourceItemSchema()
            items_list = []
            for item in items:
                items_list.append(dsi_schema.dump(item)[0])
        response = json.dumps(items_list)
        self.set_status(200)
        self.write(response)


class SetupHandler(BaseHandler):

    SUPPORTED_METHODS = ["POST", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.set_status(204)
        self.finish()

    @staticmethod
    def make_url(type_):
        """
        Make URL depending on type_
        :param type_:
        :return:
        """
        if type_ == "rss":
            netloc = Config.IMPORTER_RSS_API_HOST + ":" + str(Config.IMPORTER_RSS_API_PORT)
            tuple_ = ("http", netloc, "setup", "", "")
            return urlunsplit(tuple_)
        if type_ == "atom":
            netloc = Config.IMPORTER_ATOM_API_HOST + ":" + str(Config.IMPORTER_ATOM_API_PORT)
            tuple_ = ("http", netloc, "setup", "", "")
            return urlunsplit(tuple_)
        return None

    def post(self):
        """Load data from url.
        ---
        description: Load data from url.
        parameters:
        -   name: type
            in: body
            description: type
            required: true
            type: string
        responses:
            200:
                description: Success                
                schema:
                    type: object
                    properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        data:
                            type: object
                            description: data                                 
                        errorCode:
                            type: integer
                            description: errorCode
        """
        body = {}
        try:
            body = json_decode(self.request.body)
        except Exception as e:
            logger.error("Error getting request body: ", e)
            self.set_status(400)
            response = {"statusCode": 400, "description": "Error getting request body: {}".format(e),
                        "errorCode": None}
            response = {"response": response}
            self.write(response)
        if "type" in body:
            url = self.make_url(body["type"])

            response = requests.post(url, json=body)
            response_data = escape.json_decode(response.text)
            self.set_status(response.status_code)
            response = {"statusCode": response.status_code, "data": response_data,
                        "errorCode": None}
            response = {"response": response}
            self.write(response)
        else:
            logger.error("Missing 'type' field.")
            self.set_status(400)
            response = {"statusCode": 400, "description": "Missing 'type' field.",
                        "errorCode": None}
            response = {"response": response}
            self.write(response)


class FilesHandler(FileRequestHandler):

    SUPPORTED_METHODS = ["POST", "OPTIONS"]

    CSV_header = {'Url', 'Content', 'Type', 'Author', 'AuthorLink', 'Date', 'AggregatedSocialReactionsCount',
               'CommentsCount', 'AggregatedRepostsCount', 'Clicks', 'WikiViews', 'ViewsCount'}

    INTEGRUM_header = OrderedDict.fromkeys(['Источник', 'Дата', 'Время', 'Заголовок', 'Текст', 'Оригинал', 'Автор'])
    SOCIALINTEGRUM_header = OrderedDict.fromkeys(['Блогер', 'Дата', 'Заголовок/Текст', 'Платформа', 'Аудиторный охват'])

    Record = namedtuple('Record', 'Источник Дата Время Заголовок Текст Оригинал Автор Ошибки')
    Record_social = namedtuple('Record_social', 'Блогер Дата Заголовок_Текст Платформа Аудиторный_охват Ошибки')

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.set_status(204)
        self.finish()

    @staticmethod
    def iri_to_uri(iri):
        parts = urlsplit(iri)
        uri = urlunsplit((
            parts.scheme,
            parts.netloc.encode('idna').decode('ascii'),
            quote(parts.path),
            quote(parts.query, '='),
            quote(parts.fragment),
        ))
        return uri

    @staticmethod
    def make_url(filename):
        """
        Make URL depending on type_
        :param type_:
        :return:
        """
        netloc = Config.IMPORTER_API_HOST + ":" + str(Config.IMPORTER_API_PORT)
        tuple_ = ("http", netloc, "api/v1/static/{}".format(filename), "", "")
        return urlunsplit(tuple_)

    @staticmethod
    def csv_dict_reader(file_obj):

        """
        Read a CSV file using csv.DictReader
        """

        reader = csv.DictReader(file_obj, delimiter=';')
        rows = []
        err_list = []
        for line in reader:
            schema = SocialNetworkCSVSchema()
            result = schema.load(line)
            if len(result[1]) > 0:
                logger.error("Errors in CSV file: {}".format(result[1]))
                for k, v in result[1].items():
                    if k in line:
                        result[0][k] = line[k]
                source = result[0]
                errs = result[1]
                source["Errors"] = str(errs)
                source_errs = source
                err_list.append(source_errs)
            else:
                rows.append(schema.dump(result[0]))
        tmp_list = []
        for el in rows:
            tmp_list.append(el[0])
        pure_list = list(unique_everseen(tmp_list))
        d = {"rows": pure_list, "errors": err_list}
        return d

    @staticmethod
    def xlsx_reader(file_obj):
        """
        Read a XLSX file using openpyxl
        """
        wb = load_workbook(filename=file_obj)
        ws = wb.active
        t = tuple(ws.rows)
        header = []
        for item in t[0]:  # get firstline data (header)
            if item.value:
                header.append(item.value)
        header.append("Ошибки")
        data_rows = t[1:]
        err_list = []
        norm_list = []
        # link = re.compile(r"url=(http.+)", re.I)
        schema = IntegrumSchema()

        for row in data_rows:
            d = {}
            d['Source'] = row[0].value
            d['Date'] = row[1].value
            d['Time'] = row[2].value
            d['Title'] = row[3].value
            d['Content'] = row[4].value
            d['Uri'] = row[5].value
            d['Author'] = row[6].value

            result = schema.load(d)

            if len(result[1]) > 0:
                for k, v in result[1].items():
                    result[0][k] = d[k]  # add error data to dict w/o errors
                source = result[0]  # cleaned dict
                errs = result[1]  # dict with errors
                for k, v in errs.items():  # replace keys to cyrillic keys and delete old ones
                    if k == "Source":
                        errs["Источник"] = errs["Source"]
                        del errs["Source"]
                    if k == "Date":
                        errs["Дата"] = errs["Date"]
                        del errs["Date"]
                    if k == "Time":
                        errs["Время"] = errs["Time"]
                        del errs["Time"]
                    if k == "Title":
                        errs["Заголовок"] = errs["Title"]
                        del errs["Title"]
                    if k == "Content":
                        errs["Текст"] = errs["Content"]
                        del errs["Content"]
                    if k == "Uri":
                        errs["Оригинал"] = errs["Uri"]
                        del errs["Uri"]
                    if k == "Author":
                        errs["Автор"] = errs["Author"]
                        del errs["Author"]
                    if k == "Errors":
                        errs["Ошибки"] = errs["Errors"]
                        del errs["Errors"]
                source["Errors"] = str(errs)  # stringify all errors list
                source_errs = source
                err_list.append(source_errs)
            else:
                norm_list.append(schema.dump(result[0]))
        tmp_list = []
        for el in norm_list:
            tmp_list.append(el[0])
        pure_list = list(unique_everseen(tmp_list))
        d = {"rows": pure_list, "errors": err_list, "header": header}
        return d

    @staticmethod
    def social_xlsx_reader(file_obj):
        """
        Read a XLSX file using openpyxl
        """
        wb = load_workbook(filename=file_obj)
        ws = wb.active
        t = tuple(ws.rows)
        header = []
        for item in t[0]:  # get firstline data (header)
            if item.value:
                header.append(item.value)
        header.append("Ошибки")

        data_rows = t[1:]
        err_list = []
        norm_list = []
        link = re.compile(r"url=(http.+)", re.I)
        soc_schema = SocialIntegrumSchema()
        for row in data_rows:
            d = {}
            d['bloger'] = row[0].value
            d['dateTime'] = row[1].value
            d['headerText'] = row[2].value
            d["platform"] = row[3].value
            d['audience'] = row[4].value
            result = soc_schema.load(d)
            if len(result[1]) > 0:
                for k, v in result[1].items():
                    result[0][k] = d[k]  # add error data to dict w/o errors
                source = result[0]  # cleaned dict
                errs = result[1]  # dict with errors
                for k, v in errs.items():  # replace keys to cyrillic keys and delete old ones
                    if k == "bloger":
                        errs["Блогер"] = errs["bloger"]
                        del errs["bloger"]
                    if k == "dateTime":
                        errs["Дата"] = errs["dateTime"]
                        del errs["dateTime"]
                    if k == "headerText":
                        errs["Заголовок_текст"] = errs["headerText"]
                        del errs["headerText"]
                    if k == "platform":
                        errs["Платформа"] = errs["platform"]
                        del errs["platform"]
                    if k == "audience":
                        errs["Аудиторный_охват"] = errs["audience"]
                        del errs["audience"]
                    if k == "Errors":
                        errs["Ошибки"] = errs["Errors"]
                        del errs["Errors"]
                source["Errors"] = str(errs)  # stringify all errors list
                source_errs = source
                err_list.append(source_errs)
            else:
                norm_list.append(soc_schema.dump(result[0]))
        tmp_list = []
        for el in norm_list:
            tmp_list.append(el[0])
        pure_list = list(unique_everseen(tmp_list))
        d = {"rows": pure_list, "errors": err_list, "header": header}
        return d

    def post(self):
        """Upload files.
        ---
        description: Upload files.
        parameters:
        -   name: type
            in: body
            description: type
            required: true
            type: string
        -   name: file
            in: body
            description: file
            required: true
            type: file
        responses:
            200:
                description: Success
        """

        systemTimeZoneId = self.get_argument("systemTimeZoneId", default="Europe/Moscow")
        if systemTimeZoneId == "":
            systemTimeZoneId = "Europe/Moscow"
        type = self.get_argument("type", default="ContentCSV")

        file = self.request.files['file'][0]
        filename = file['filename']
        extension = os.path.splitext(filename)[1]
        name = translit(filename, "ru", reversed=True)
        if len(filename) > 255:
            logger.error("File name is too long")
            self.set_status(400)
            self.set_header("Error-description", "Filename is too long.")
            return
        if type == "ContentCSV" and extension == ".csv":

            """
            Encoding check
            """
            try:
                content = file["body"].decode('utf-8-sig')
            except Exception as E:
                logger.error("Wrong file encoding (must be UTF-8)")
                self.set_status(400)
                self.set_header("Error-description", "Wrong file encoding (must be UTF-8)")
                return

            uploaded_file = io.StringIO()
            uploaded_file.write(content)

            """
            Check correct delimiter(we have ';') and header
            """
            reader = csv.reader(uploaded_file.getvalue().splitlines())
            headers = next(reader, None)
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(headers[0])
            if dialect.delimiter != ';':
                logger.error("Wrong delimiter. Must be ';', You have '{}'".format(dialect.delimiter))
                self.set_status(400)
                self.set_header("Error-description", "Wrong delimiter. Must be ';', You have '{}'".format(dialect.delimiter))
                return

            """header check"""

            header = headers[0].split(';')
            header = set(header)
            if not self.CSV_header.issubset(header):
                logger.error("Wrong headers in CSV file")
                self.set_status(400)
                self.set_header("Error-Description", "Wrong headers in CSV file")
                return

            """
            Continue file handling
            """

            result = self.csv_dict_reader(uploaded_file.getvalue().splitlines())
            if len(result["errors"]) > 0:
                """
                Implement errors writing to CSV here
                """
                inmemoryfile = io.StringIO()
                w = csv.DictWriter(inmemoryfile, result["errors"][0].keys(), delimiter=";")
                w.writeheader()
                w.writerows(result["errors"])

            """
            Process successful rows
            """
            if result:
                with filestore.atomic() as tx:
                    for row in result["rows"]:
                        date = datetime.strptime(row["Date"], '%H:%M %d.%m.%Y')
                        try:
                            SocialNetworkCSV.create(
                                id=uuid4(),
                                Url=row["Url"].strip(),
                                Content=row["Content"],
                                Type=row["Type"],
                                Author=row["Author"],
                                AuthorLink=row["AuthorLink"],
                                Date=date,
                                AggregatedSocialReactionsCount=row["AggregatedSocialReactionsCount"],
                                CommentsCount=row["CommentsCount"],
                                AggregatedRepostsCount=row["AggregatedRepostsCount"],
                                Clicks=row["Clicks"],
                                WikiViews=row["WikiViews"],
                                ViewsCount=row["ViewsCount"],
                                is_read=False,
                                filename=filename,
                                timezone=systemTimeZoneId
                                )
                        except Exception as ex:
                            tx.rollback()
                            logger.error("Error creating SocialNetworkCSV object: {}".format(ex))
                            self.set_status(400)
                            self.set_header("Error-Description", "Error creting SocialNetworkCSV object. Please try again. Possible problem with DataBase connection")
                            return
                if len(result["errors"]) > 0:
                    self.set_status(200)
                    self.set_header("Content-Type", "text/csv")
                    self.set_header("Content-Disposition", "attachment; filename={}".format(name.replace(" ", "")))
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    self.set_header("Rows-failed", "{}".format(len(result["errors"])))
                    self.write(inmemoryfile.getvalue())
                else:
                    self.set_status(200)
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    logger.info("Uploaded {}, {} rows".format(name.replace(" ", ""), len(result["rows"])))

            """
            Integrum format (SMI)
            """
        elif type == "Integrum" and extension == ".xlsx":
            content = file["body"]
            uploaded_file = io.BytesIO()
            uploaded_file.write(content)

            """
            Check header
            """
            wb = load_workbook(filename=uploaded_file)
            ws = wb.active
            t = tuple(ws.rows)
            header = []
            for item in t[0]:  # get firstline data (header)
                if item.value:
                    header.append(item.value)
            header = OrderedDict.fromkeys(header)
            if not self.INTEGRUM_header == header:
                logger.error("Wrong headers in Integrum file")
                self.set_status(400)
                self.set_header("Error-Description",
                                "Wrong headers in Integrum file")
                return

            result = self.xlsx_reader(uploaded_file)

            err_tuple_list = []
            """
            Process successful rows
            """
            if result:
                with filestore.atomic() as tx1:
                    for row in result["rows"]:
                        datetime_ = row["Time"] + " " + row["Date"]
                        date = datetime.strptime(datetime_, '%H:%M %d.%m.%y')
                        try:
                            Integrum.create(
                                id=uuid4(),
                                Source=row["Source"],
                                Date=row["Date"],
                                Time=row["Time"],
                                DateTime=date,
                                Title=row["Title"],
                                Content=row["Content"],
                                Uri=row["Uri"].strip(),
                                Author=row["Author"],
                                is_read=False,
                                filename=filename,
                                timezone=systemTimeZoneId
                                )
                        except Exception as Exc:
                            tx1.rollback()
                            logger.error("Error creting Integrum object: {}".format(Exc))
                            self.set_status(400)
                            self.set_header("Error-Description",
                                            "Error creting Integrum object. Please try again. Possible problem with DataBase connection")
                            return
                if len(result["errors"]) > 0:
                    """
                    Implement errors writing to xlsx here
                    """
                    memoryfile = io.BytesIO()
                    for item in result["errors"]:
                        err_tuple = self.Record(item["Source"], item["Date"], item["Time"], item["Title"], item["Content"],
                                           item["Uri"], item["Author"], item["Errors"])
                        err_tuple_list.append(err_tuple)

                    wb_errors = Workbook()
                    err_sheet = wb_errors.active
                    err_sheet.title = "Тематическая карта"
                    err_sheet.append(result["header"])

                    for row in err_tuple_list:
                        err_sheet.append(row)
                    wb_errors.save(memoryfile)

                    name = translit(filename, "ru", reversed=True)
                    self.set_status(200)
                    self.set_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.set_header("Content-Disposition", "attachment; filename={}".format(name.replace(" ", "")))
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    self.set_header("Rows-failed", "{}".format(len(result["errors"])))
                    self.write(memoryfile.getvalue())
                else:
                    self.set_status(200)
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    logger.info("Uploaded {}, {} rows".format(name.replace(" ", ""), len(result["rows"])))

        elif type == "SocialIntegrum" and extension == ".xlsx":
            content = file["body"]
            social_int_file = io.BytesIO()
            social_int_file.write(content)

            """
            Check header
            """
            wb = load_workbook(filename=social_int_file)
            ws = wb.active
            t = tuple(ws.rows)
            header = []
            for item in t[0]:  # get firstline data (header)
                if item.value:
                    header.append(item.value)
            header = OrderedDict.fromkeys(header)
            if not self.SOCIALINTEGRUM_header == header:
                logger.error("Wrong headers in SocialIntegrum file")
                self.set_status(400)
                self.set_header("Error-Description", "Wrong headers in SocialIntegrum file")
                return

            result = self.social_xlsx_reader(social_int_file)

            err_tuple_list = []
            """
            Process successful rows
            """
            if result:
                with filestore.atomic() as tx2:
                    for row in result["rows"]:
                        date = datetime.strptime(row["dateTime"], '%d.%m.%y %H:%M')
                        try:
                            SocialIntegrum.create(
                                id=uuid4(),
                                bloger=row["bloger"],
                                dateTime=date,
                                headerText=row["headerText"],
                                platform=row["platform"].strip(),
                                audience=row["audience"],
                                is_read=False,
                                filename=filename,
                                timezone=systemTimeZoneId
                            )
                        except Exception as Except:
                            tx2.rollback()
                            logger.error("Error creating SocialIntegrum object: {}".format(Except))
                            self.set_status(400)
                            self.set_header("Error-Description", "Error creating SocialIntegrum object. Please try again. Possible "
                                                       "problem with DataBase connection")
                            return

                if len(result["errors"]) > 0:
                    """
                    Implement errors writing to xlsx here
                    """
                    memfile = io.BytesIO()
                    for item in result["errors"]:
                        errors_tuple = self.Record_social(item["bloger"], item["dateTime"], item["headerText"],
                                                       item["platform"], item["audience"], item["Errors"])
                        err_tuple_list.append(errors_tuple)

                    wb_errors = Workbook()
                    err_sheet = wb_errors.active
                    err_sheet.title = "Тематическая карта"
                    err_sheet.append(result["header"])

                    for row in err_tuple_list:
                        err_sheet.append(row)
                    wb_errors.save(memfile)
                    name = translit(filename, "ru", reversed=True)
                    self.set_status(200)
                    self.set_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.set_header("Content-Disposition", "attachment; filename={}".format(name.replace(" ", "")))
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    self.set_header("Rows-failed", "{}".format(len(result["errors"])))
                    self.write(memfile.getvalue())
                else:
                    self.set_status(200)
                    self.set_header("Rows-successful", "{}".format(len(result["rows"])))
                    logger.info("Uploaded {}, {} rows".format(name.replace(" ", ""), len(result["rows"])))
        else:
            logger.error("Error: File type is not supported")
            self.set_status(400)
            self.set_header("Error-Description", "File type is not supported")


class DataSourceItemStatsHandler(BaseHandler):

    SUPPORTED_METHODS = ["GET", "OPTIONS"]

    def options(self):
        self.set_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.set_status(204)
        self.finish()

    def get(self, datasource_id, item_id):
        """Returns info about DataSourceItems Stats.
        ---
        description: Returns info about DataSourceItems Stats. 
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        statusCode:
                            type: integer
                            description: statusCode
                        description:
                            type: string
                            description: count                              
                        errorCode:
                            type: string
                            description: errorCode
        """
        self.set_status(200)
        response = {"statusCode": 200, "description": "Some stats will be here",
                    "errorCode": None}
        response = {"response": response}
        self.write(response)


class ContentHandler(tornado.web.RequestHandler):

    """Mocked Content Api"""

    SUPPORTED_METHODS = ["GET"]

    def get(self):
        if self.get_argument("filters"):
            filters = self.get_query_argument("filters")
            list_ = []
            d = {
                    "uid": "413498dd-fd67-4503-be20-c30af7430cab",
                    "url": "https://www.facebook.com/100000181265285/posts/2635522426463774",
                    "title": "t",
                    "content": "c",
                    "dataSourceItemUid": "43a716f6-5d38-11e9-bb13-3b6e1feb509f",
                    "importedDateTime": "2019-04-12T15:35:41.429378",
                    "publishedDate": None,
                    "publishedTime": None,
                    "publishedDateTime": "2019-04-12T15:35:41.429378",
                    "viewsCount": None,
                    "aggregatedSocialReactionsCount": None,
                    "aggregatedRepostsCount": None,
                    "commentsCount": None,
                    "followersCount": None,
                    "createdDate": "2019-04-12T15:36:29.730886",
                    "modifiedDate": "2019-04-12T15:36:29.730886"
                  }
            list_.append(d)
            response = {"items": list_}
            self.write(response)

class GeoIpHandler(tornado.web.RequestHandler):

    """Get geoIp information"""

    SUPPORTED_METHODS = ["POST"]

    def post(self):
        """Get geoIp information
        ---
        description: Get geoIp information
        parameters:
        -   name: urls
            in: body
            description: urls
            required: true
            type: string        
        responses:
            200:
                description: Success                
                schema:
                   type: object
                   properties:
                        geo:
                            type: string
                            description: geo                       
                        timezone:
                            type: string
                            description: timezone
        """
        raw_body = None
        try:
            raw_body = json_decode(self.request.body)
        except Exception as e:
            logger.error(("Error reading POST body: {}".format(e)))
            self.set_status(400)
            logger.error("Error in request body: {}".format(e))
            response = {"statusCode": 400, "description": "Error {}".format(e)}
            self.write(response)
            return
        try:
            if raw_body:
                params = {'apiKey': Config.GEOIP_API_KEY}
                result_list = []
                for url in raw_body['urls']:
                    split_url = urlsplit(url)
                    ip = socket.gethostbyname(split_url.netloc)
                    params['ip'] = ip
                    outer_response = requests.get('https://api.ipgeolocation.io/timezone', params=params)
                    data = outer_response.json()
                    dict_to_send = data['geo']
                    dict_to_send['timezone'] = data['timezone']
                    result_list.append(dict_to_send)
                response = json.dumps(result_list)
                logger.info("GeoIp found, result is {} ".format(response))
                self.write(response)
        except Exception as e:
            logger.error(("Error receiving GeoIP: {}".format(e)))
            self.set_status(400)
            logger.error("Error receiving GeoIP: {}".format(e))
            response = {"statusCode": 400, "description": "Error {}".format(e)}
            self.write(response)